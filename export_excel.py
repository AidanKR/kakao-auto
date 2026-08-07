"""
엑셀(xlsx) 내보내기 — 실무는 결국 엑셀. 방목록·응답대기·메시지를 한 파일로.

시트:
  1) 방목록   : 방, 누적건수, 최초, 마지막, 참여자수
  2) 응답대기 : 방, 미답(시간), 질문여부, 마지막 메시지 (briefing 규칙 재사용)
  3) 메시지   : 시각, 방, 보낸이, 내용 (최신 N건, --msgs 로 조절/생략)

openpyxl 필요(1_install.bat 이 설치). 한글 헤더, 헤더 고정+필터.
LLM·외부 전송 없음(전부 로컬).

사용: python export_excel.py                 (share_dir/카카오톡.xlsx)
      python export_excel.py --msgs 20000    (메시지 시트 최신 2만건)
      python export_excel.py --no-msgs       (메시지 시트 생략)
"""
import argparse
from datetime import datetime
from pathlib import Path

import db
import briefing

HERE = Path(__file__).parent


def _load_cfg():
    return briefing.load_config()


def _autosize(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header(ws, cols):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(cols)
    fill = PatternFill("solid", fgColor="2A78D6")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:" + ws.cell(row=1, column=len(cols)).coordinate


def build(out_path, cfg, msgs_limit):
    try:
        import openpyxl  # noqa
    except ImportError:
        print("openpyxl 미설치 — 'pip install openpyxl' (또는 1_install.bat) 후 재시도")
        return False
    from openpyxl import Workbook

    now = datetime.now()
    conn = db.connect()
    wb = Workbook()

    # 1) 방목록
    ws = wb.active
    ws.title = "방목록"
    _header(ws, ["방", "누적건수", "최초", "마지막", "참여자수"])
    members = {}
    for room, sender in conn.execute(
        "SELECT DISTINCT room, sender FROM messages WHERE sender!=''"
    ).fetchall():
        members.setdefault(room, set()).add(sender)
    for room, first, last, cnt in db.all_rooms(conn):
        ws.append([room, cnt, (first or "")[:16].replace("T", " "),
                   (last or "")[:16].replace("T", " "), len(members.get(room, ()))])
    _autosize(ws, [30, 10, 18, 18, 9])

    # 2) 응답대기
    ws2 = wb.create_sheet("응답대기")
    _header(ws2, ["방", "미답(시간)", "질문", "연속", "마지막 메시지"])
    me = briefing.my_names(cfg, conn)
    for p in briefing.triage(conn, cfg, me, now):
        last = p["burst"][-1][2].replace("\n", " ") if p["burst"] else ""
        ws2.append([p["room"], round(p["age_h"], 1) if p["age_h"] is not None else "",
                    "예" if p["question"] else "", p["count"], last[:120]])
    _autosize(ws2, [26, 10, 6, 6, 60])

    # 3) 메시지
    if msgs_limit and msgs_limit > 0:
        ws3 = wb.create_sheet("메시지")
        _header(ws3, ["시각", "방", "보낸이", "내용"])
        rows = conn.execute(
            "SELECT sent_at, room, sender, content FROM messages "
            "ORDER BY sent_at DESC LIMIT ?", (msgs_limit,)
        ).fetchall()
        for sa, rm, sn, ct in rows:
            ws3.append([(sa or "")[:16].replace("T", " "), rm, sn or "",
                        (ct or "").replace("\n", " ")[:400]])
        _autosize(ws3, [17, 24, 14, 80])
        print(f"  메시지 시트: {len(rows):,}건")

    conn.close()
    wb.save(out_path)
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", help="출력 xlsx 경로")
    ap.add_argument("--msgs", type=int, default=20000, help="메시지 시트 최신 N건(기본 20000)")
    ap.add_argument("--no-msgs", action="store_true", help="메시지 시트 생략")
    args = ap.parse_args()

    cfg = _load_cfg()
    limit = 0 if args.no_msgs else args.msgs
    if args.out:
        out = Path(args.out)
    else:
        base = Path(cfg.get("share_dir") or (HERE / "share")).expanduser()
        base.mkdir(parents=True, exist_ok=True)
        out = base / "카카오톡.xlsx"

    if build(out, cfg, limit):
        print(f"엑셀 생성: {out}")


if __name__ == "__main__":
    main()
